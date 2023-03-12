import os  # obrigatorio para limpar o terminal
from datetime import date, datetime # biblioteca para obter data/hora
from openpyxl import Workbook, load_workbook

# Carregar o modelo na variavel
wb = load_workbook('modelo/modOrd.xlsx')
# Escolhe active sheet
ws = wb.active

# salva e formata a data
# datetime.today() -> salva a data completa
# .strftime() -> formada a tada de acordo com os argumentos
data = datetime.today().strftime("%d-%m-%Y")

# --- inicio class info ---
class info:

    # --- inicio __init__ ---
    def __init__(self, adm):
        self.adm = 'sam'

        self.cliente = ""
        self.mecanico = ""
        self.data = ""

        self.numMaquina = ""
        self.nomeMaquina = ""
        self.modMaquina = ""

        self.compra = ""
        self.reembolso = ""

        self.numPeca = ""
        self.descProduto = ""
        self.qnteProduto = ""

        self.caminhoArquivo = 'arquivos/'
    # --- fim __init__ ---

    # --- inicio infoNovaOrd ---
    def infoNovaOrd(self):
        novaOrd = input('\nNova ordem de serviço? (s/n): ')
        # limpa a tela
        os.system("cls")

        if novaOrd == 's':

            reembolso = input('\nÉ um Reembolso? (s/n): ')
            # limpa a tela
            os.system("cls")

            if reembolso == 's':
                # caso seja reembolso
                self.reembolso = '(X)'

                compra = input('\nComprar peças? (s/n): ')
                # limpa a tela
                os.system("cls")

                if compra == 's':
                    self.compra = '(X)'
                else:
                    self.compra = ' '
                # --- fim if ---

                # nome do cliente
                cliente = input('\nNome do Cliente: ')
                # limpa a tela
                os.system("cls")

                self.cliente = cliente

            else:
                # caso não seja reembolso
                self.reembolso = ' '
                self.compra = '(X)'
            # --- fim if ---

            nomeMecanico = input('\nMecanico Responsavel: ')
            self.mecanico = nomeMecanico
            # limpa a tela
            os.system("cls")

            # salva a data
            self.data = data

            maquina = input('\nNome da Maquina: ')
            self.nomeMaquina = maquina
            # limpa a tela
            os.system("cls")

            modelo = input('\nModelo da Maquina: ')
            self.modMaquina = modelo
            # limpa a tela
            os.system("cls")

            numeroMaquina = input('\nNumero da Maquina: ')
            self.numMaquina = numeroMaquina
            # limpa a tela
            os.system("cls")

            ordem = input('\nGerar Nova Ordem de Serviço? (s/n): ')
            # limpa a tela
            os.system("cls")
            if ordem == 's':
                self.gerarOrd()
            else:
                print('\nAté Mais')

        else:
            print('\naté mais')
        # --- fim if ---
    # --- fim infoNovaOrd ---

    # --- inicio gerarOrd ---
    def gerarOrd(self):

        # insere as informações salvas nas celulas especificadas
        ws['B3'] = self.cliente
        ws['E3'] = self.mecanico
        ws['B5'] = self.data
        ws['B10'] = self.nomeMaquina
        ws['B12'] = self.numMaquina
        ws['E10'] = self.modMaquina
        ws['B14'] = self.reembolso
        ws['B15'] = self.compra

        # cria o nome pra o arquivo
        nomeArquivo = self.caminhoArquivo+self.nomeMaquina+'_'+self.modMaquina+'_'+self.numMaquina+'.xlsx'
        # salva o arquvio com o nom novo
        wb.save(nomeArquivo)
        print('\nArquivo salvo')
    # --- fim gerarOrd ---

    def admin(self):
        self.adm

# --- fim class info ---
